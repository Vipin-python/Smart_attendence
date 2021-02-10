'''
smart attendence
version: 5.0.1
Copyright © 2021  - by vip 
'''

#   ''pip install -r requirements.txt'' 

import xlrd
import xlwt
import getpass 
import openpyxl
import datetime
from xlutils.copy import copy
from xlrd import open_workbook
from urllib.request import urlopen
from prettytable import PrettyTable
from openpyxl.styles import PatternFill
pt=PrettyTable()
link = "https://httxlzqgrt1uoe8yoo8axg-on.drv.tw/Webpage/pass.html"
f = urlopen(link)
myfile = f.read()
x=str(myfile)
int=x[18:21]
today=datetime.datetime.now()
print(' DATE-',today.strftime("%x"))
print(' TIME-',today.strftime("%X"))
print(' ')
# for passward by getpass if wrong passward enter loop will be continue running
p = getpass.getpass(prompt='Your Name Sir? ')
user = getpass.getuser()
while True:
	   pwd = getpass.getpass("User Pass : %s" % user)
	   if pwd == int:
	       print(' ')
	       print ("Welcome !!!")
	       print(' ')
	       break
	   else:
	   	print ("The password you entered is incorrect.")
	   	
print('************************************************************')
pt.field_names=['HELLO',p.upper(),'SIR']
pt.add_row([' ','GOOD MORNING , WELCOME TO SMART ATTENDENCE',' '])
print(pt)
print(' ')

# it take file name if any error found it again ask for file
print(' This will accept excel file which are present only in the folder of this executable file of python')
print(' ')
print(' You can run excel file from any other folder by giving file path with extenstion with a "\\\\" ')
print(' ')
print(' You can use file excel file with extenstions -\n  .xls ,  .xlsx  ,  .xlsm  ,  .xlts  ,  .xltm')
print(' ')
while True:
	loc =str(input("Enter your file name: "))
	try:
		#this try to open xls file of excel with the prettytable
		p=PrettyTable()
		wb = xlrd.open_workbook(loc)
		sheet =wb.sheet_by_index(0)
		sheet.cell_value(0,0)
		x=sheet.nrows
		p.field_names=sheet.row_values(0)
		for i in range(1,x):
			p.add_row(sheet.row_values(i))
		print(p)
		print(' ')
		y=int(input('Enter column no. of Names: '))
		c=sheet.col_values(y-1)
		col=sheet.col_values(0)
		break
	except:
			try:
				#this try to read your xlsx,xlsm,xltx and xltm files
				t=list()
				c=list()
				col=list()
				wb_obj = openpyxl.load_workbook(loc)
				sheet_obj = wb_obj.active
				m_col =sheet_obj.max_column+1
				m_row=sheet_obj.max_row
				
				def iter_rows(sheet_obj,a):
					result=list()
					for row in sheet_obj.iter_rows(a,a):
						for cell in row:
							result.append(cell.value)
							yield result
				z=list(iter_rows(sheet_obj,1))
				tab=PrettyTable()
				tab.field_names=z[0]
				no=m_row+1				
				for i in range(2,no):
					z=list(iter_rows(sheet_obj,i))
					tab.add_row(z[0])
				print(tab)
								
				def path(path,y,nam):
					wb_obj = openpyxl.load_workbook(path)
					sheet_obj = wb_obj.active
					for i in range(1, m_col):
						cell_obj = sheet_obj.cell(row = y, column = i)
						t.append(cell_obj.value)
					for i in range(1,m_row+1):
						cell_obj1=sheet_obj.cell(row=i,column=nam)
						c.append(cell_obj1.value)
				row1=int(input('Enter your column no of Roll no: '))
				nam=int(input('Enter column no. of Names: '))

				for i in range(1,m_col):
					path(loc,i,nam)
					for i in range(1,m_row+1):
						cell_obj2=sheet_obj.cell(row=i,column=row1)
						col.append(cell_obj2.value)
						x=len(col)-m_row
						while x>0:
							col.pop(0)
							c.pop(0)
							x=x-1								
				
				
				break
			except:
				print("file is not found try again")

#this function try to make letter into title form so that we can easily compared				
def cap(c):
		x=len(c)
		for i in range(x):
			a=c[i].title()
			c.append(a)
		while x>0:
			c.pop(0)
			x=x-1
			
cap(c)
print(' ')
print(c[0],end='- ')
print(c[1:])
print(col[0],end='- ')
print(col[1:])
print(' ')
l=list()

count=0

enter='Enter your Name or Roll no: '
Q=str(input('Do you want to make excel of attendence(yes or no): '))

print('|**** To exit type exit() this command**** |')

if Q=='yes':
	print(' ')
	enter='Enter your Roll no: '
	print(' ')
	file=str(input('Enter your excel file name with or .xlsx to save it: '))
	try:
	       	excel_file = openpyxl.load_workbook(loc)
	       	excel_sheet = excel_file.active
	       	excel_sheet.cell(row=1, column=m_col).value ='Attendence'
	       	for i in range(1,m_row):
	       		excel_sheet.cell(row=i+1, column=m_col).value ='Absent'
	       		excel_sheet.cell(row=i+1, column=m_col).fill=PatternFill('solid',fgColor='DC143C')
	       		excel_file.save(file)

	except:
		rb = open_workbook(loc)
		wb = copy(rb)
		s = wb.get_sheet(0)
		s.write(0,x-1,'Attendence')
		for i in range(0,x-1):
			s.write(i+1,x-1,'absent')
			wb.save(file)

        	
while True:
	print('\n  Present',count,'   Absent',len(col)-1-count,'     Total=',len(col)-1)
	
	print(' ')
	
	a=input(enter).title()
	print(' ')
	if l.count(a)==0:
		try:
			if a=='Exit()':
				break
			elif c[1:].count(a)>0 or col[1:].count(int(a))>0 :
				l.append(a)
				n=dict(zip(col,c))
				try:
					a=int(a)
					print(' ',n[a],'is mark present')
					count=count+1
				except:
					a=str(a)
					print(' ',a,'is mark present')
					count=count+1
				
				if Q=='yes':
					a=int(a)
					try:
							excel_sheet.cell(row=a+1, column=m_col).value ='Present'
							excel_sheet.cell(row=a+1, column=m_col).fill=PatternFill('solid',fgColor='00FF00')
							excel_file.save(file)
					except:
						s.write(a,x-1,'present')
						wb.save(file)
							
			else:
				print('\nNot found')
		except:
			print('\nError found ')
	else:
		print('\nalready marked present')
print('Thank you have a nice day')